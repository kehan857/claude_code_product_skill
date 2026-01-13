#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "自媒体平台注册汇总"

# 定义表头
headers = ["平台名称", "个人注册链接", "个人注册说明", "企业注册链接", "企业注册说明"]
ws.append(headers)

# 定义数据
data = [
    [
        "微信公众号",
        "https://mp.weixin.qq.com/",
        "订阅号，需要身份证及人脸识别验证，每天可发1条消息",
        "https://mp.weixin.qq.com/",
        "服务号，需要营业执照及对公账户验证，每月可发4条消息，支持支付功能"
    ],
    [
        "头条号",
        "https://mp.toutiao.com/",
        "需要身份证实名认证，年满18周岁",
        "https://mp.toutiao.com/",
        "需要营业执照，支持蓝V认证，获得平台流量倾斜"
    ],
    [
        "百家号",
        "https://baijiahao.baidu.com/",
        "需要百度账号及身份证实名认证",
        "https://baijiahao.baidu.com/",
        "需要营业执照，企业权威认证费用600元/年"
    ],
    [
        "腾讯新闻（企鹅号）",
        "https://om.qq.com/",
        "可用QQ或微信账号注册，审核时间1-3个工作日",
        "https://om.qq.com/",
        "支持企业、媒体、政府等多种类型，需要相应资质证明"
    ],
    [
        "知乎",
        "https://www.zhihu.com/",
        "手机号即可注册，无需其他认证",
        "https://www.zhihu.com/org/signup",
        "机构号目前免费，需企业资质认证，享有蓝色官方认证标识"
    ],
    [
        "视频号",
        "https://channels.weixin.qq.com/",
        "需要微信账号登录，一个微信号对应一个视频号",
        "https://channels.weixin.qq.com/",
        "企业视频号需要企业微信认证，享有更多功能权限"
    ],
    [
        "小红书",
        "https://www.xiaohongshu.com/",
        "手机号注册，可申请专业号，免费",
        "https://business.xiaohongshu.com/",
        "企业号认证费用300元/年，需营业执照"
    ],
    [
        "抖音",
        "https://creator.douyin.com/",
        "手机号注册，需实名认证（身份证）",
        "https://renzheng.douyin.com/",
        "企业蓝V认证费用600元/年，需营业执照、对公账户等"
    ]
]

# 添加数据行
for row in data:
    ws.append(row)

# 设置列宽
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 45

# 设置表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# 设置数据行样式
data_font = Font(name="微软雅黑", size=11)
data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
link_font = Font(name="微软雅黑", size=11, color="0563C1", underline="single")

# 边框样式
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

# 应用样式到所有单元格
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
    for cell in row:
        cell.border = thin_border
        if cell.row > 1:  # 数据行
            cell.font = data_font
            cell.alignment = data_alignment
            # 为链接列添加超链接样式
            if cell.column in [2, 4]:  # B列和D列是链接列
                cell.font = link_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # 添加超链接
                if cell.value:
                    cell.hyperlink = cell.value

# 设置行高
ws.row_dimensions[1].height = 30
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 60

# 添加筛选器
ws.auto_filter.ref = f"A1:E{ws.max_row}"

# 冻结首行
ws.freeze_panes = "A2"

# 保存文件
output_path = "/Users/kexiaobin/Desktop/其他/claude code/自媒体平台注册汇总.xlsx"
wb.save(output_path)
print(f"Excel文件已创建: {output_path}")
