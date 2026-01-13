#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "VC专职人员报备"

# 定义表头
headers = [
    "序号",
    "VC合作伙伴名称（与PRM主体一致）",
    "填写人姓名",
    "填写人角色",
    "VC专职人员姓名",
    "VC人员岗位角色",
    "人员类型",
    "业务范围",
    "简历文件名",
    "提交日期"
]

ws.append(headers)

# 设置列宽
column_widths = {
    'A': 8,   # 序号
    'B': 30,  # VC合作伙伴名称
    'C': 15,  # 填写人姓名
    'D': 20,  # 填写人角色
    'E': 15,  # VC专职人员姓名
    'F': 15,  # VC人员岗位角色
    'G': 25,  # 人员类型
    'H': 20,  # 业务范围
    'I': 25,  # 简历文件名
    'J': 15   # 提交日期
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 设置表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 边框样式
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

# 应用表头样式
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

ws.row_dimensions[1].height = 30

# 添加示例数据行（可选）
example_data = [
    ["1", "", "", "伙伴老板/管理层", "", "面销", "在职人员，在职时间3个月以上", "只做火山业务", "", ""],
    ["2", "", "", "伙伴操盘手", "", "电销", "新招募入职", "火山+其他业务", "", ""],
]

for row_data in example_data:
    ws.append(row_data)

# 设置数据行样式
data_font = Font(name="微软雅黑", size=10)
data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
    for cell in row:
        cell.border = thin_border
        cell.font = data_font
        cell.alignment = data_alignment
        if cell.row == 2:
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[2].height = 25
ws.row_dimensions[3].height = 25

# 添加数据验证（下拉菜单）
# D列：填写人角色
dv_d = DataValidation(
    type="list",
    formula1='"伙伴老板/管理层,伙伴操盘手,其他"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='输入错误',
    error='请从下拉列表中选择',
    errorStyle='warning'
)
dv_d.prompt = '请选择填写人角色'
dv_d.promptTitle = '填写人角色'
ws.add_data_validation(dv_d)
dv_d.add('D2:D1000')

# F列：VC人员岗位角色
dv_f = DataValidation(
    type="list",
    formula1='"面销,电销,架构师,技术,其他"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='输入错误',
    error='请从下拉列表中选择',
    errorStyle='warning'
)
dv_f.prompt = '请选择岗位角色'
dv_f.promptTitle = 'VC人员岗位角色'
ws.add_data_validation(dv_f)
dv_f.add('F2:F1000')

# G列：人员类型
dv_g = DataValidation(
    type="list",
    formula1='"在职人员，在职时间3个月以上,新招募入职"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='输入错误',
    error='请从下拉列表中选择',
    errorStyle='warning'
)
dv_g.prompt = '请选择人员类型'
dv_g.promptTitle = '人员类型'
ws.add_data_validation(dv_g)
dv_g.add('G2:G1000')

# H列：业务范围
dv_h = DataValidation(
    type="list",
    formula1='"只做火山业务,火山+其他业务"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle='输入错误',
    error='请从下拉列表中选择',
    errorStyle='warning'
)
dv_h.prompt = '请选择业务范围'
dv_h.promptTitle = '业务范围'
ws.add_data_validation(dv_h)
dv_h.add('H2:H1000')

# 冻结首行
ws.freeze_panes = "A2"

# 添加说明工作表
ws_info = wb.create_sheet("填写说明")
info_content = [
    ["2026年VC专职人员报备表单 - 填写说明", ""],
    ["", ""],
    ["表格说明：", ""],
    ["1. 每位VC专职人员需要单独填写一行", ""],
    ["2. 标有 * 的字段为必填项", ""],
    ["", ""],
    ["字段说明：", ""],
    ["序号", "自动编号，方便统计"],
    ["VC合作伙伴名称", "请填写与PRM主体一致的合作伙伴名称"],
    ["填写人姓名", "填写本表单的人员姓名"],
    ["填写人角色", "请从下拉列表选择：伙伴老板/管理层、伙伴操盘手、其他"],
    ["VC专职人员姓名", "需要报备的专职人员姓名"],
    ["VC人员岗位角色", "请从下拉列表选择：面销、电销、架构师、技术、其他"],
    ["人员类型", "请从下拉列表选择：在职人员（3个月以上）、新招募入职"],
    ["业务范围", "请从下拉列表选择：只做火山业务、火山+其他业务"],
    ["简历文件名", "上传简历后，填写文件名"],
    ["提交日期", "填写提交表单的日期"],
    ["", ""],
    ["注意事项：", ""],
    ["• 请确保填写信息准确无误", ""],
    ["• 简历文件请单独保存，统一命名格式：姓名-VC专职人员简历", ""],
    ["• 建议定期备份Excel数据", ""],
]

for row_data in info_content:
    ws_info.append(row_data)

# 设置说明工作表的样式
ws_info.column_dimensions['A'].width = 30
ws_info.column_dimensions['B'].width = 50

title_cell = ws_info['A1']
title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="4472C4")
ws_info.merge_cells('A1:B1')

for row in ws_info.iter_rows(min_row=1, max_row=len(info_content), min_col=1, max_col=2):
    for cell in row:
        if cell.row == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.column == 1 and "说明" in str(cell.value) and "注意" not in str(cell.value):
                cell.font = Font(name="微软雅黑", size=11, bold=True, color="4472C4")
            elif "注意" in str(cell.value):
                cell.font = Font(name="微软雅黑", size=11, bold=True, color="C00000")

ws_info.row_dimensions[1].height = 30

# 保存文件
output_path = "/Users/kexiaobin/Desktop/其他/claude code/2026年VC专职人员报备表.xlsx"
wb.save(output_path)
print(f"✅ Excel文件已创建: {output_path}")
print("\n📋 表单字段：")
print("1. VC合作伙伴名称（文本）")
print("2. 填写人姓名（文本）")
print("3. 填写人角色（下拉选择：伙伴老板/管理层、伙伴操盘手、其他）")
print("4. VC专职人员姓名（文本）")
print("5. VC人员岗位角色（下拉选择：面销、电销、架构师、技术、其他）")
print("6. 人员类型（下拉选择：在职人员、新招募入职）")
print("7. 业务范围（下拉选择：只做火山业务、火山+其他业务）")
print("8. 简历文件名（文本）")
print("9. 提交日期（日期）")
